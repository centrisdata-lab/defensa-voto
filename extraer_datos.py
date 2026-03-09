import openpyxl, json, unicodedata

def normalize(s):
    s = str(s).strip().upper()
    s = unicodedata.normalize('NFD', s)
    s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
    return s

wb = openpyxl.load_workbook(
    'c:/Users/sara.munoz/OneDrive - arus.com.co/Analitica EO/AnalisisDescriptivo_Info/Lideres/Defensa del voto/TESTIGOS ESCRUTADORES.xlsx'
)

data = {}

# ---- OTROS MUNICIPIOS (col0=municipio, col1=comision texto, col2=nombre) ----
ws = wb['OTROS MUNICIPIOS']
for row in ws.iter_rows(min_row=2, values_only=True):
    municipio = row[0]; comision = row[1]; nombre = row[2]
    if not municipio or not comision or not nombre:
        continue
    municipio = normalize(municipio)
    comision_str = normalize(comision)
    nombre_str = normalize(nombre)

    if municipio not in data:
        data[municipio] = {'total_personas': 0, 'municipales': 0, 'auxiliares': 0,
                           'comisiones_unicas': set(), 'personas': []}
    data[municipio]['total_personas'] += 1
    if 'MUNICIPAL' in comision_str:
        data[municipio]['municipales'] += 1
    if 'AUXILIAR' in comision_str:
        data[municipio]['auxiliares'] += 1
    data[municipio]['comisiones_unicas'].add(comision_str)
    data[municipio]['personas'].append({'nombre': nombre_str, 'comision': comision_str})

# ---- MEDELLIN 2 (col0=comision, col1=nombre) ----
# Comisiones: número = auxiliar, "Municipal"/"Municipal remante" = municipal, "Departamental N" = departamental
ws2 = wb['MEDELLIN 2']
municipio = 'MEDELLIN'
data[municipio] = {'total_personas': 0, 'municipales': 0, 'auxiliares': 0, 'departamentales': 0,
                   'comisiones_unicas': set(), 'personas': [], 'auxiliares_sin_nombre': []}
for row in ws2.iter_rows(min_row=2, values_only=True):
    com_raw = row[0]; nombre = row[1]
    # Registrar auxiliares sin nombre
    if not nombre and com_raw is not None:
        com_str = str(com_raw).strip()
        if com_str.isdigit():
            data[municipio]['auxiliares_sin_nombre'].append(f'AUXILIAR {com_raw}')
        continue
    if not nombre:
        continue
    nombre_str = normalize(nombre)
    com_str = normalize(str(com_raw)) if com_raw is not None else ''

    # Clasificar tipo de comisión
    if 'MUNICIPAL' in com_str:
        tipo = 'municipal'
        comision_str = com_str
    elif 'DEPARTAMENTAL' in com_str:
        tipo = 'departamental'
        comision_str = com_str
    else:
        # Número = comisión auxiliar
        tipo = 'auxiliar'
        comision_str = f'AUXILIAR {com_raw}'

    data[municipio]['total_personas'] += 1
    if tipo == 'municipal':
        data[municipio]['municipales'] += 1
    elif tipo == 'auxiliar':
        data[municipio]['auxiliares'] += 1
    elif tipo == 'departamental':
        data[municipio]['departamentales'] += 1
    data[municipio]['comisiones_unicas'].add(comision_str)
    data[municipio]['personas'].append({'nombre': nombre_str, 'comision': comision_str})

# Serializar
output = {}
for mun, d in data.items():
    output[mun] = {
        'total_personas': d['total_personas'],
        'municipales': d['municipales'],
        'auxiliares': d['auxiliares'],
        'departamentales': d.get('departamentales', 0),
        'auxiliares_sin_nombre': sorted(d.get('auxiliares_sin_nombre', []), key=lambda x: int(x.split()[-1]) if x.split()[-1].isdigit() else 999),
        'total_comisiones': len(d['comisiones_unicas']),
        'comisiones': sorted(list(d['comisiones_unicas'])),
        'personas': d['personas']
    }

with open(
    'c:/Users/sara.munoz/OneDrive - arus.com.co/Analitica EO/AnalisisDescriptivo_Info/Lideres/Defensa del voto/escrutadores_data.json',
    'w', encoding='utf-8'
) as f:
    json.dump(output, f, ensure_ascii=False, indent=2)

total = sum(d['total_personas'] for d in output.values())
print(f'OK - Municipios: {len(output)}, Total personas: {total}')
for mun, d in sorted(output.items()):
    print(f'  {mun}: {d["total_personas"]} personas, {d["municipales"]} municipales, {d["auxiliares"]} auxiliares, {d["departamentales"]} departamentales')
