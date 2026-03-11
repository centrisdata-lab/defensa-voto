import pandas as pd
import json
import openpyxl

BASE_PATH = "c:/Users/sara.munoz/OneDrive - arus.com.co/Analitica EO/AnalisisDescriptivo_Info/Lideres/Defensa del voto/analisis/"
API_URL = "https://estrategia-electoral.onrender.com"

# ---- Cargar DIVIPOLA para cruzar comisiones faltantes ----
wb_div = openpyxl.load_workbook(BASE_PATH + "DIVIPOLA.xlsx", read_only=True)
ws_div = wb_div.active
div_rows = list(ws_div.iter_rows(values_only=True))
# Columnas: DEPARTAMENTO=0, MUNICIPIO=1, ZONA=2, PUESTO=3, dep=4, mun=5, nom=6, tipo_com=7, nom_com=8
# Construir lookup: (municipio_upper, zona_zfill2, puesto_zfill2) -> nom_com
divipola_lookup = {}
for r in div_rows[1:]:
    dep_nom = r[4]
    if str(dep_nom).upper() != 'ANTIOQUIA':
        continue
    mun = str(r[5]).upper().strip() if r[5] else ''
    zona = str(r[2]).zfill(2) if r[2] is not None else '00'
    puesto = str(r[3]).zfill(2) if r[3] is not None else '00'
    nom_com = r[8] or ''
    key = (mun, zona, puesto)
    if key not in divipola_lookup:
        divipola_lookup[key] = str(nom_com)

# ---- Cargar Excel principal: todos los municipios, sin filtro ----
wb = openpyxl.load_workbook(BASE_PATH + "Análisis mesa Departamento Antioquia (1) (1).xlsx.xlsx", read_only=True)
ws = wb.active
all_rows = list(ws.iter_rows(values_only=True))
col_headers = all_rows[0]
# Indices: Departamento=0, Municipio=1, Zona=2, ID Puesto=3, Comisión=4,
#          Nombre Puesto=5, Mesa #=6, Escrutado=7, Ahora Colombia=8,
#          Partido Conservador=9, Votos Blanco=10, Votos Nulos=11,
#          Votos No Marcados=12, Total Votos Mesa=13, Prioridad Mesa=14,
#          Votos Senado Dif 10%=15, Votos Cámara Dif 10%=16, Dif 10%=17,
#          Duplicidad=18, Acción=19, Medio=20

def get_comision(row):
    """Retorna la comisión: del Excel si existe, sino cruza con DIVIPOLA."""
    com = str(row[4]).strip() if row[4] else ''
    if com:
        return com
    mun = str(row[1]).upper().strip() if row[1] else ''
    zona = str(row[2]).zfill(2) if row[2] else '00'
    puesto = str(row[3]).zfill(2) if row[3] else '00'
    key = (mun, zona, puesto)
    found = divipola_lookup.get(key, '')
    if found:
        return found
    # Si no se encuentra con zona+puesto, buscar solo por municipio (nivel municipal)
    key_mun = (mun, '00', '00')
    return divipola_lookup.get(key_mun, '(Sin comisión)')

def safe_int(val):
    try:
        return int(float(val)) if val not in (None, '', 'None') else 0
    except:
        return 0

def safe_float(val):
    try:
        return float(val) if val not in (None, '', 'None') else 0.0
    except:
        return 0.0

def safe_str(val):
    return str(val).strip() if val not in (None, '', 'None') else ''

# Agrupar por municipio -> lista de mesas (sin ningún filtro)
data2 = {}
totals2 = {}

for row in all_rows[1:]:
    mun = safe_str(row[1])
    if not mun:
        continue

    comision = get_comision(row)
    nombre_puesto = safe_str(row[5])
    mesa = safe_str(row[6])
    zona = safe_str(row[2])
    ahora = safe_int(row[8])
    conservador = safe_int(row[9])
    votos_blanco = safe_int(row[10])
    votos_nulos = safe_int(row[11])
    votos_no_marcados = safe_int(row[12])
    total_votos = safe_int(row[13])
    prioridad = safe_str(row[14])
    votos_senado_dif = safe_str(row[15])
    votos_camara_dif = safe_str(row[16])
    dif_10pct = safe_str(row[17])
    duplicidad = safe_str(row[18])
    accion = safe_str(row[19])
    medio = safe_str(row[20])
    diferencia = conservador - ahora

    entry = {
        'zona': zona,
        'comision': comision,
        'nombre_puesto': nombre_puesto,
        'mesa': mesa,
        'votos_ahora': ahora,
        'votos_conservador': conservador,
        'votos_blanco': votos_blanco,
        'votos_nulos': votos_nulos,
        'votos_no_marcados': votos_no_marcados,
        'total_votos': total_votos,
        'diferencia': diferencia,
        'prioridad': prioridad,
        'votos_senado_dif': votos_senado_dif,
        'votos_camara_dif': votos_camara_dif,
        'dif_10pct': dif_10pct,
        'duplicidad': duplicidad,
        'accion': accion,
        'medio': medio,
    }

    if mun not in data2:
        data2[mun] = []
    data2[mun].append(entry)

# Calcular totales por municipio
for mun, rows in data2.items():
    totals2[mun] = {
        'total_ahora': sum(r['votos_ahora'] for r in rows),
        'total_conservador': sum(r['votos_conservador'] for r in rows),
        'diferencia_total': sum(r['diferencia'] for r in rows),
        'total_mesas': len(rows),
    }

MUNICIPIOS = sorted(data2.keys())

print(f"Municipios: {len(MUNICIPIOS)}")
print(f"Total mesas: {sum(len(v) for v in data2.values())}")
sin_com = sum(1 for rows in data2.values() for r in rows if r['comision'] == '(Sin comisión)')
print(f"Sin comisión asignada: {sin_com}")

data2_json = json.dumps(data2, ensure_ascii=False)
totals_json = json.dumps(totals2, ensure_ascii=False)
municipios_json = json.dumps(MUNICIPIOS, ensure_ascii=False)
api_url_json = json.dumps(API_URL)

html = """<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Defensa del Voto — Antioquia</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap" rel="stylesheet">
<style>
  :root {
    --primary: #001361;
    --primary-light: #1e40af;
    --accent: #06b6d4;
    --success: #10b981;
    --success-light: #d1fae5;
    --warning: #f59e0b;
    --warning-light: #fef3c7;
    --danger: #ef4444;
    --danger-light: #fee2e2;
    --dark: #1e293b;
    --gray: #64748b;
    --light: #f1f5f9;
    --white: #ffffff;
    --gradient: linear-gradient(135deg, #001361 0%, #1e40af 100%);
    --shadow: 0 4px 6px -1px rgba(0,0,0,0.1), 0 2px 4px -1px rgba(0,0,0,0.06);
    --shadow-lg: 0 10px 15px -3px rgba(0,0,0,0.1), 0 4px 6px -2px rgba(0,0,0,0.05);
  }
  * { margin:0; padding:0; box-sizing:border-box; }
  body { font-family:'Inter',sans-serif; background:linear-gradient(180deg,#f8fafc 0%,#e2e8f0 100%); color:var(--dark); min-height:100vh; line-height:1.6; }

  .header { background:var(--gradient); color:white; padding:0.75rem 1.5rem; position:sticky; top:0; z-index:100; box-shadow:var(--shadow-lg); }
  .header-content { max-width:1400px; margin:0 auto; display:flex; justify-content:space-between; align-items:center; }
  .header h1 { font-size:1.1rem; font-weight:700; display:flex; align-items:center; gap:0.5rem; }
  .header-subtitle { font-size:0.7rem; opacity:0.9; margin-top:0.15rem; }
  .header-badge { background:rgba(255,255,255,0.15); padding:0.35rem 0.75rem; border-radius:6px; font-size:0.7rem; backdrop-filter:blur(10px); }

  .main-container { max-width:1400px; margin:0 auto; padding:1.5rem; }

  .step-card { background:var(--white); border-radius:16px; padding:2rem; box-shadow:var(--shadow-lg); margin-bottom:1.5rem; }
  .step-label { font-size:0.7rem; font-weight:700; color:var(--accent); text-transform:uppercase; letter-spacing:1px; margin-bottom:0.5rem; }
  .step-title { font-size:1.3rem; font-weight:800; color:var(--primary); margin-bottom:1.25rem; }
  .mun-grid { display:flex; flex-wrap:wrap; gap:0.6rem; }
  .mun-btn { padding:0.55rem 1.1rem; border-radius:8px; border:2px solid #e2e8f0; background:var(--light); color:var(--dark); font-family:'Inter',sans-serif; font-size:0.82rem; font-weight:600; cursor:pointer; transition:all 0.2s; }
  .mun-btn:hover { border-color:var(--primary-light); background:#eff6ff; color:var(--primary); }
  .mun-btn.active { background:var(--gradient); color:white; border-color:var(--primary); box-shadow:var(--shadow); }

  .content-panel { display:none; }
  .content-panel.visible { display:block; }

  .filters-row { background:var(--white); border-radius:10px; padding:0.85rem 1rem; margin-bottom:1rem; display:flex; gap:0.75rem; align-items:center; flex-wrap:wrap; box-shadow:var(--shadow); }
  .filters-row label { font-size:0.75rem; font-weight:700; color:var(--primary); white-space:nowrap; }
  .filters-row input, .filters-row select { background:var(--light); border:1px solid #e2e8f0; color:var(--dark); border-radius:6px; padding:0.4rem 0.75rem; font-size:0.82rem; font-family:'Inter',sans-serif; min-width:160px; }
  .filters-row input:focus, .filters-row select:focus { outline:none; border-color:var(--primary-light); }

  .summary-grid { display:flex; gap:0.75rem; margin-bottom:1rem; flex-wrap:wrap; }
  .summary-card { background:var(--white); border-radius:12px; padding:0.75rem 1rem; box-shadow:var(--shadow); display:flex; align-items:center; gap:0.75rem; flex:1; min-width:180px; position:relative; overflow:hidden; }
  .summary-card::before { content:''; position:absolute; top:0; left:0; bottom:0; width:4px; }
  .summary-card.primary::before { background:var(--gradient); }
  .summary-card.success::before { background:var(--success); }
  .summary-card.warning::before { background:var(--warning); }
  .summary-card.danger::before { background:var(--danger); }
  .summary-card .sc-icon { width:36px; height:36px; border-radius:8px; display:flex; align-items:center; justify-content:center; font-size:1rem; flex-shrink:0; }
  .summary-card.primary .sc-icon { background:#dbeafe; }
  .summary-card.success .sc-icon { background:var(--success-light); }
  .summary-card.warning .sc-icon { background:var(--warning-light); }
  .summary-card.danger .sc-icon { background:var(--danger-light); }
  .summary-card .sc-label { font-size:0.68rem; color:var(--gray); font-weight:500; text-transform:uppercase; letter-spacing:0.3px; }
  .summary-card .sc-val { font-size:1.25rem; font-weight:800; color:var(--dark); line-height:1.2; }

  .table-section { background:var(--white); border-radius:12px; padding:1rem; box-shadow:var(--shadow); overflow:hidden; margin-bottom:1rem; }
  .table-section-title { font-size:0.88rem; font-weight:700; color:var(--dark); margin-bottom:0.75rem; display:flex; align-items:center; gap:0.4rem; }
  .count-pill { background:var(--gradient); color:white; border-radius:20px; padding:0.1rem 0.6rem; font-size:0.7rem; font-weight:700; }
  .table-wrap { overflow-x:auto; max-height:520px; overflow-y:auto; }
  table { width:100%; border-collapse:collapse; font-size:0.8rem; }
  thead { position:sticky; top:0; z-index:2; }
  th { background:var(--gradient); color:white; padding:0.6rem 0.65rem; text-align:left; font-weight:600; font-size:0.75rem; white-space:nowrap; }
  th:first-child { border-radius:6px 0 0 0; } th:last-child { border-radius:0 6px 0 0; }
  td { padding:0.5rem 0.65rem; border-bottom:1px solid #e2e8f0; color:var(--dark); }
  tr:hover td { background:var(--light); }
  tr:nth-child(even) td { background:#f8fafc; }
  tr:nth-child(even):hover td { background:var(--light); }
  .badge { display:inline-flex; align-items:center; padding:0.2rem 0.55rem; border-radius:12px; font-size:0.7rem; font-weight:600; }
  .badge-alta { background:#fee2e2; color:#dc2626; }
  .badge-media { background:#fef3c7; color:#d97706; }
  .badge-baja { background:#d1fae5; color:#059669; }
  .badge-ok { background:var(--success-light); color:#059669; }
  .diff-pos { color:#dc2626; font-weight:700; }
  .diff-neg { color:#059669; font-weight:600; }
  .no-data { text-align:center; padding:2.5rem; color:var(--gray); font-size:0.9rem; }

  /* PDF link button */
  .pdf-btn { display:inline-flex; align-items:center; justify-content:center; width:26px; height:26px; border-radius:6px; background:#eff6ff; border:1px solid #bfdbfe; color:var(--primary-light); cursor:pointer; text-decoration:none; font-size:0.85rem; transition:all 0.15s; }
  .pdf-btn:hover { background:var(--primary); color:white; border-color:var(--primary); }

  /* DRILLDOWN */
  .dd-comision { background:var(--white); border:1px solid #e2e8f0; border-radius:10px; margin-bottom:8px; overflow:hidden; box-shadow:var(--shadow); }
  .dd-com-header { display:flex; align-items:center; gap:0.6rem; padding:0.75rem 1rem; cursor:pointer; user-select:none; background:linear-gradient(135deg,#eff6ff,#dbeafe); transition:background 0.2s; }
  .dd-com-header:hover { background:linear-gradient(135deg,#dbeafe,#bfdbfe); }
  .dd-com-arrow { font-size:0.75rem; color:var(--primary); transition:transform 0.2s; flex-shrink:0; }
  .dd-com-arrow.open { transform:rotate(90deg); }
  .dd-com-title { font-weight:700; color:var(--primary); font-size:0.88rem; flex:1; }
  .dd-totals { display:flex; gap:1rem; flex-wrap:wrap; }
  .dd-totals span { font-size:0.72rem; color:var(--gray); }
  .dd-totals span b { color:var(--dark); font-weight:700; }
  .dd-com-body { display:none; padding:0 0.75rem 0.75rem; background:#f8fafc; }
  .dd-com-body.open { display:block; }
  .dd-puesto { background:var(--white); border:1px solid #e2e8f0; border-radius:8px; margin-bottom:6px; overflow:hidden; }
  .dd-pue-header { display:flex; align-items:center; gap:0.6rem; padding:0.6rem 0.85rem; cursor:pointer; user-select:none; background:#f0fdf4; transition:background 0.2s; }
  .dd-pue-header:hover { background:#dcfce7; }
  .dd-pue-arrow { font-size:0.7rem; color:var(--success); transition:transform 0.2s; flex-shrink:0; }
  .dd-pue-arrow.open { transform:rotate(90deg); }
  .dd-pue-title { font-weight:600; color:#15803d; font-size:0.82rem; flex:1; }
  .dd-pue-body { display:none; padding:0 0.75rem 0.75rem; }
  .dd-pue-body.open { display:block; }
  .dd-mesa-table { width:100%; border-collapse:collapse; font-size:0.78rem; margin-top:4px; }
  .dd-mesa-table th { background:var(--gradient); color:white; padding:0.45rem 0.65rem; font-size:0.72rem; font-weight:600; text-align:left; white-space:nowrap; }
  .dd-mesa-table td { padding:0.4rem 0.65rem; border-bottom:1px solid #e2e8f0; color:var(--dark); }
  .dd-mesa-table tr:last-child td { border-bottom:none; }
  .dd-mesa-table tr:hover td { background:var(--light); }
  .dd-total-row td { background:#eff6ff !important; font-weight:700; color:var(--primary) !important; border-top:2px solid var(--primary-light); }
  .row-alta td { background:#fff7ed !important; }
  .row-alta:hover td { background:#ffedd5 !important; }

  /* Modal PDF */
  .modal-overlay { display:none; position:fixed; inset:0; background:rgba(0,0,0,0.6); z-index:1000; align-items:center; justify-content:center; }
  .modal-overlay.open { display:flex; }
  .modal-box { background:white; border-radius:12px; width:92vw; height:90vh; display:flex; flex-direction:column; overflow:hidden; box-shadow:0 25px 50px rgba(0,0,0,0.3); }
  .modal-header { background:var(--gradient); color:white; padding:0.75rem 1rem; display:flex; justify-content:space-between; align-items:center; flex-shrink:0; }
  .modal-header h3 { font-size:0.88rem; font-weight:700; }
  .modal-close { background:rgba(255,255,255,0.2); border:none; color:white; width:28px; height:28px; border-radius:6px; cursor:pointer; font-size:1rem; display:flex; align-items:center; justify-content:center; }
  .modal-close:hover { background:rgba(255,255,255,0.35); }
  .modal-body { flex:1; overflow:hidden; }
  .modal-body iframe { width:100%; height:100%; border:none; }

  /* Comentarios */
  .com-section { margin-top:6px; padding:6px 0 2px; border-top:1px dashed #e2e8f0; }
  .com-list { margin-bottom:4px; }
  .com-item { background:#f8fafc; border-radius:6px; padding:4px 8px; margin-bottom:3px; font-size:0.72rem; }
  .com-item-nombre { font-weight:700; color:var(--primary); margin-right:6px; }
  .com-item-fecha { color:var(--gray); font-size:0.65rem; margin-right:6px; }
  .com-item-texto { color:var(--dark); margin-top:2px; }
  .com-add-btn { background:#eff6ff; border:1px solid #bfdbfe; color:var(--primary-light); border-radius:6px; padding:0.2rem 0.6rem; font-size:0.7rem; cursor:pointer; font-family:'Inter',sans-serif; }
  .com-add-btn:hover { background:var(--primary); color:white; }
  .com-form { display:none; gap:4px; flex-wrap:wrap; margin-top:4px; align-items:center; }
  .com-form input, .com-form textarea { border:1px solid #e2e8f0; border-radius:6px; padding:0.3rem 0.5rem; font-size:0.72rem; font-family:'Inter',sans-serif; background:white; }
  .com-form textarea { min-width:200px; resize:vertical; }
  .com-form button { background:var(--primary); color:white; border:none; border-radius:6px; padding:0.3rem 0.7rem; font-size:0.72rem; cursor:pointer; font-family:'Inter',sans-serif; }
  .com-form button.cancel { background:#e2e8f0; color:var(--dark); }
  .com-del-btn { background:none; border:none; color:#ef4444; cursor:pointer; font-size:0.75rem; margin-left:4px; padding:0 2px; }

  @media(max-width:640px){
    .main-container{padding:1rem;}
    .mun-grid{gap:0.4rem;}
    .mun-btn{font-size:0.75rem; padding:0.45rem 0.8rem;}
  }
</style>
</head>
<body>

<header class="header">
  <div class="header-content">
    <div>
      <h1>&#x1F5F3;&#xFE0F; Defensa del Voto &mdash; Antioquia</h1>
      <div class="header-subtitle">An&aacute;lisis electoral &middot; Congreso 2026</div>
    </div>
    <div class="header-badge">MIRA &middot; InfoMIRA</div>
  </div>
</header>

<!-- Modal PDF -->
<div class="modal-overlay" id="pdf-modal" onclick="closeModal(event)">
  <div class="modal-box">
    <div class="modal-header">
      <h3 id="modal-title">Acta de Mesa</h3>
      <button class="modal-close" onclick="closePdfModal()">&#x2715;</button>
    </div>
    <div class="modal-body">
      <iframe id="pdf-iframe" src=""></iframe>
    </div>
  </div>
</div>

<div class="main-container">

  <div class="step-card">
    <div class="step-label">Paso 1</div>
    <div class="step-title">Seleccione un municipio</div>
    <div style="margin-bottom:1rem;">
      <input type="text" id="mun-search" placeholder="&#x1F50D; Buscar municipio..." oninput="filterMunicipios()"
        style="width:100%;max-width:400px;padding:0.55rem 1rem;border-radius:8px;border:2px solid #e2e8f0;font-family:'Inter',sans-serif;font-size:0.9rem;outline:none;transition:border-color 0.2s;"
        onfocus="this.style.borderColor='#1e40af'" onblur="this.style.borderColor='#e2e8f0'">
    </div>
    <div class="mun-grid" id="mun-grid"></div>
  </div>

  <div id="step2" style="display:none">
    <!-- Filtros -->
    <div class="filters-row">
      <label>Comisi&oacute;n:</label>
      <input type="text" id="fc-comision" placeholder="Buscar comisi&oacute;n..." oninput="renderMesas()" style="min-width:150px;">
      <label>Puesto:</label>
      <input type="text" id="fc-puesto" placeholder="Buscar puesto..." oninput="renderMesas()">
      <label>Mesa:</label>
      <input type="text" id="fc-mesa" placeholder="Nro mesa..." oninput="renderMesas()" style="min-width:100px;">
      <label>Prioridad:</label>
      <select id="fc-prioridad" onchange="renderMesas()">
        <option value="">Todas</option>
        <option value="ALTA">ALTA</option>
        <option value="MEDIA">MEDIA</option>
        <option value="BAJA">BAJA</option>
      </select>
    </div>

    <div class="summary-grid" id="mun-stats"></div>
    <div id="mesas-drill"></div>
  </div>

</div>

<script>
const DATA2 = """ + data2_json + """;
const TOTALS = """ + totals_json + """;
const MUNICIPIOS = """ + municipios_json + """;
const API_URL = """ + api_url_json + """;

let currentMun = null;

// Ping cada 4 min para mantener la API despierta en Render free
setInterval(() => { fetch(API_URL + '/').catch(() => {}); }, 4 * 60 * 1000);

function safeId(s) { return String(s).replace(/[^a-zA-Z0-9]/g, '_'); }

function buildPdfUrl(municipio, zona, puesto, mesa) {
  const z = String(zona).padStart(2, '0');
  const m = String(mesa).padStart(3, '0');
  return API_URL + '/CAMARA/ANTIOQUIA/' + encodeURIComponent(municipio) + '/Zona%20' + z + '/' + encodeURIComponent(puesto) + '/mesa_' + m + '.pdf';
}

function openPdfModal(url, label) {
  document.getElementById('modal-title').textContent = label;
  document.getElementById('pdf-iframe').src = url;
  document.getElementById('pdf-modal').classList.add('open');
}
function closePdfModal() {
  document.getElementById('pdf-modal').classList.remove('open');
  document.getElementById('pdf-iframe').src = '';
}
function closeModal(e) { if (e.target === document.getElementById('pdf-modal')) closePdfModal(); }

(function init(){
  const grid = document.getElementById('mun-grid');
  MUNICIPIOS.forEach(m => {
    const btn = document.createElement('button');
    btn.className = 'mun-btn';
    btn.textContent = m;
    btn.onclick = () => selectMun(m);
    grid.appendChild(btn);
  });
})();

function filterMunicipios() {
  const q = document.getElementById('mun-search').value.toLowerCase().trim();
  document.querySelectorAll('.mun-btn').forEach(btn => {
    btn.style.display = btn.textContent.toLowerCase().includes(q) ? '' : 'none';
  });
}

function selectMun(mun) {
  currentMun = mun;
  document.querySelectorAll('.mun-btn').forEach(b => b.classList.toggle('active', b.textContent === mun));
  document.getElementById('step2').style.display = 'block';
  ['fc-comision','fc-puesto','fc-mesa'].forEach(id => document.getElementById(id).value = '');
  document.getElementById('fc-prioridad').value = '';
  renderMesas();
  document.getElementById('step2').scrollIntoView({behavior:'smooth', block:'start'});
}

function renderMesas() {
  if (!currentMun) return;
  const fCom      = document.getElementById('fc-comision').value.toLowerCase().trim();
  const fPuesto   = document.getElementById('fc-puesto').value.toLowerCase().trim();
  const fMesa     = document.getElementById('fc-mesa').value.toLowerCase().trim();
  const fPrioridad = document.getElementById('fc-prioridad').value.toUpperCase().trim();

  const totals = TOTALS[currentMun] || {};
  let rows = DATA2[currentMun] || [];
  if (fCom)      rows = rows.filter(r => r.comision && r.comision.toLowerCase().includes(fCom));
  if (fPuesto)   rows = rows.filter(r => r.nombre_puesto && r.nombre_puesto.toLowerCase().includes(fPuesto));
  if (fMesa)     rows = rows.filter(r => r.mesa && r.mesa.toString().includes(fMesa));
  if (fPrioridad) rows = rows.filter(r => r.prioridad && r.prioridad.toUpperCase() === fPrioridad);

  const altaCount = rows.filter(r => r.prioridad && r.prioridad.toUpperCase() === 'ALTA').length;
  const mediaCount = rows.filter(r => r.prioridad && r.prioridad.toUpperCase() === 'MEDIA').length;

  document.getElementById('mun-stats').innerHTML = `
    <div class="summary-card primary"><div class="sc-icon">&#x1F5F3;&#xFE0F;</div><div><div class="sc-label">Total mesas</div><div class="sc-val">${rows.length}</div></div></div>
    <div class="summary-card danger"><div class="sc-icon">&#x26A0;&#xFE0F;</div><div><div class="sc-label">Prioridad ALTA</div><div class="sc-val">${altaCount}</div></div></div>
    <div class="summary-card warning"><div class="sc-icon">&#x1F4CA;</div><div><div class="sc-label">Prioridad MEDIA</div><div class="sc-val">${mediaCount}</div></div></div>
    <div class="summary-card success"><div class="sc-icon">&#x1F3DB;&#xFE0F;</div><div><div class="sc-label">Total Ahora Col.</div><div class="sc-val">${(totals.total_ahora||0).toLocaleString()}</div></div></div>
    <div class="summary-card warning"><div class="sc-icon">&#x1F534;</div><div><div class="sc-label">Total Conservador</div><div class="sc-val">${(totals.total_conservador||0).toLocaleString()}</div></div></div>
  `;

  const drill = document.getElementById('mesas-drill');
  if (!rows.length) {
    drill.innerHTML = '<div class="no-data">No se encontraron mesas para los filtros aplicados en ' + currentMun + '</div>';
    return;
  }

  // Agrupar por comision -> puesto
  const byComision = {};
  rows.forEach(r => {
    const com = r.comision || '(Sin comisi\u00F3n)';
    const pue = r.nombre_puesto || '(Sin puesto)';
    if (!byComision[com]) byComision[com] = {};
    if (!byComision[com][pue]) byComision[com][pue] = [];
    byComision[com][pue].push(r);
  });

  let html = `<div class="table-section"><div class="table-section-title">Agrupado por Comisi\u00F3n \u2192 Puesto \u2192 Mesa <span class="count-pill">${rows.length} mesas</span></div><div>`;

  Object.keys(byComision).sort().forEach((com, ci) => {
    const puestos = byComision[com];
    const allRows = Object.values(puestos).flat();
    const comAhora = allRows.reduce((s,r) => s+(r.votos_ahora||0),0);
    const comCons  = allRows.reduce((s,r) => s+(r.votos_conservador||0),0);
    const comDif   = allRows.reduce((s,r) => s+(r.diferencia||0),0);
    const comAlta  = allRows.filter(r => r.prioridad && r.prioridad.toUpperCase()==='ALTA').length;
    const comId = 'com-' + ci;

    let pHtml = '';
    Object.keys(puestos).sort().forEach((pue, pi) => {
      const mesas = puestos[pue];
      const pAhora = mesas.reduce((s,r) => s+(r.votos_ahora||0),0);
      const pCons  = mesas.reduce((s,r) => s+(r.votos_conservador||0),0);
      const pDif   = mesas.reduce((s,r) => s+(r.diferencia||0),0);
      const pueId  = 'pue-' + ci + '-' + pi;

      const mesasRows = mesas.map(r => {
        const dif = r.diferencia || 0;
        const zona = r.zona || '';
        const mesa = r.mesa || '';
        const pdfUrl = buildPdfUrl(currentMun, zona, pue, mesa);
        const mesaLabel = ('Mesa ' + mesa + ' - ' + pue).replace(/'/g,"\\'");
        const rowClass = r.prioridad && r.prioridad.toUpperCase()==='ALTA' ? 'row-alta' : '';
        const priBadge = r.prioridad ? `<span class="badge badge-${r.prioridad.toLowerCase()}">${r.prioridad}</span>` : '';
        const accion = r.accion || '';
        const dif10 = r.dif_10pct || '';
        const comKey = safeId(currentMun + '_' + pue + '_' + mesa);

        return `<tr class="${rowClass}">
          <td style="text-align:center">${mesa}</td>
          <td style="text-align:right">${(r.votos_ahora||0).toLocaleString()}</td>
          <td style="text-align:right">${(r.votos_conservador||0).toLocaleString()}</td>
          <td style="text-align:right" class="${dif>0?'diff-pos':'diff-neg'}">${dif>0?'+':''}${dif.toLocaleString()}</td>
          <td style="text-align:center">${r.total_votos||0}</td>
          <td style="text-align:center">${priBadge}</td>
          <td style="text-align:center">${dif10}</td>
          <td style="font-size:0.72rem;max-width:200px">${accion}</td>
          <td style="text-align:center"><a class="pdf-btn" onclick="openPdfModal('${pdfUrl}','${mesaLabel}')" title="Ver acta PDF">&#x1F50D;</a></td>
          <td>
            <div class="com-section">
              <div class="com-list" id="com-list-${comKey}"><span style="color:#94a3b8;font-size:0.68rem">Cargando...</span></div>
              <button class="com-add-btn" id="com-add-${comKey}" onclick="toggleComForm('${comKey}',true)">+ Comentario</button>
              <div class="com-form" id="com-form-${comKey}">
                <input type="text" id="com-nombre-${comKey}" placeholder="Tu nombre" style="min-width:100px;max-width:130px">
                <textarea id="com-texto-${comKey}" placeholder="Comentario..." rows="2"></textarea>
                <button onclick="saveComentario('${comKey}','${currentMun.replace(/'/g,"\\'")}','${pue.replace(/'/g,"\\'")}','${mesa}')">Guardar</button>
                <button class="cancel" onclick="toggleComForm('${comKey}',false)">Cancelar</button>
              </div>
            </div>
          </td>
        </tr>`;
      }).join('');

      const mesaKeys = mesas.map(r => safeId(currentMun + '_' + pue + '_' + r.mesa)).join(',');
      pHtml += `<div class="dd-puesto">
        <div class="dd-pue-header" data-keys="${mesaKeys}" onclick="toggleDD('${pueId}',this.querySelector('.dd-pue-arrow'),this.dataset.keys)">
          <span class="dd-pue-arrow">&#9654;</span>
          <span class="dd-pue-title">${pue}</span>
          <div class="dd-totals">
            <span><b>${mesas.length}</b> mesas</span>
            <span>Ahora: <b>${pAhora.toLocaleString()}</b></span>
            <span>Conservador: <b>${pCons.toLocaleString()}</b></span>
            <span class="${pDif>0?'diff-pos':'diff-neg'}">Dif: <b>${pDif>0?'+':''}${pDif.toLocaleString()}</b></span>
          </div>
        </div>
        <div class="dd-pue-body" id="${pueId}">
          <table class="dd-mesa-table">
            <thead><tr>
              <th style="text-align:center">Mesa</th>
              <th style="text-align:right">Ahora Col.</th>
              <th style="text-align:right">Conservador</th>
              <th style="text-align:right">Diferencia</th>
              <th style="text-align:center">Total Votos</th>
              <th style="text-align:center">Prioridad</th>
              <th style="text-align:center">Dif 10%</th>
              <th>Acci\u00F3n</th>
              <th style="text-align:center">Acta</th>
              <th>Comentarios</th>
            </tr></thead>
            <tbody>
              ${mesasRows}
              <tr class="dd-total-row">
                <td style="text-align:center">TOTAL</td>
                <td style="text-align:right">${pAhora.toLocaleString()}</td>
                <td style="text-align:right">${pCons.toLocaleString()}</td>
                <td style="text-align:right">${pDif>0?'+':''}${pDif.toLocaleString()}</td>
                <td colspan="6"></td>
              </tr>
            </tbody>
          </table>
        </div>
      </div>`;
    });

    html += `<div class="dd-comision">
      <div class="dd-com-header" onclick="toggleDD('${comId}',this.querySelector('.dd-com-arrow'),null)">
        <span class="dd-com-arrow">&#9654;</span>
        <span class="dd-com-title">${com}</span>
        <div class="dd-totals">
          <span><b>${allRows.length}</b> mesas</span>
          <span>Ahora: <b>${comAhora.toLocaleString()}</b></span>
          <span>Conservador: <b>${comCons.toLocaleString()}</b></span>
          <span class="${comDif>0?'diff-pos':'diff-neg'}">Dif: <b>${comDif>0?'+':''}${comDif.toLocaleString()}</b></span>
          ${comAlta>0?`<span style="color:#dc2626"><b>${comAlta}</b> ALTA</span>`:''}
        </div>
      </div>
      <div class="dd-com-body" id="${comId}">${pHtml}</div>
    </div>`;
  });

  html += '</div></div>';
  drill.innerHTML = html;
}

function toggleDD(id, arrow, keysStr) {
  const body = document.getElementById(id);
  const open = body.classList.toggle('open');
  if (arrow) arrow.classList.toggle('open', open);
  // Cargar comentarios al abrir un puesto (keysStr es string separado por comas)
  if (open && keysStr) {
    keysStr.split(',').forEach(k => { if (k) loadComentarios(k); });
  }
}

// ---- Comentarios ----
function loadComentarios(key) {
  const listEl = document.getElementById('com-list-' + key);
  if (!listEl) return;
  const parts = key.split('_');
  const mesa = parts[parts.length - 1];
  const puesto = parts.slice(1, parts.length - 1).join('_');
  const municipio = parts[0];
  fetch(`${API_URL}/comentarios?municipio=${encodeURIComponent(municipio)}&puesto=${encodeURIComponent(puesto)}&mesa=${encodeURIComponent(mesa)}`)
    .then(r => r.json())
    .then(data => {
      if (!data.length) {
        listEl.innerHTML = '<span style="color:#94a3b8;font-size:0.68rem">Sin comentarios</span>';
      } else {
        listEl.innerHTML = data.map(c =>
          `<div class="com-item">
            <span class="com-item-nombre">${c.nombre}</span>
            <span class="com-item-fecha">${c.fecha}</span>
            <button class="com-del-btn" onclick="deleteComentario(${c.id},'${key}')" title="Borrar">&#x2715;</button>
            <div class="com-item-texto">${c.texto}</div>
          </div>`
        ).join('');
      }
    })
    .catch(() => {
      listEl.innerHTML = '<span style="color:#ef4444;font-size:0.68rem">Error al cargar</span>';
    });
}

function toggleComForm(key, show) {
  const form = document.getElementById('com-form-' + key);
  const btn  = document.getElementById('com-add-'  + key);
  if (!form) return;
  form.style.display = show ? 'flex' : 'none';
  btn.style.display  = show ? 'none' : 'inline-block';
}

function saveComentario(key, municipio, puesto, mesa) {
  const nombreEl = document.getElementById('com-nombre-' + key);
  const textoEl  = document.getElementById('com-texto-'  + key);
  const nombre = nombreEl ? nombreEl.value.trim() : '';
  const texto  = textoEl  ? textoEl.value.trim()  : '';
  if (!nombre || !texto) { alert('Escribe tu nombre y el comentario.'); return; }
  fetch(`${API_URL}/comentarios`, {
    method: 'POST',
    headers: {'Content-Type': 'application/json'},
    body: JSON.stringify({municipio, puesto, mesa, nombre, texto})
  })
  .then(r => r.json())
  .then(data => {
    if (data.ok) {
      if (textoEl) textoEl.value = '';
      toggleComForm(key, false);
      loadComentarios(key);
    } else {
      alert('Error: ' + (data.error || 'intenta de nuevo'));
    }
  })
  .catch(() => alert('Error de conexi\u00f3n con el servidor.'));
}

function deleteComentario(id, key) {
  if (!confirm('¿Borrar este comentario?')) return;
  fetch(`${API_URL}/comentarios/${id}`, { method: 'DELETE' })
    .then(r => r.json())
    .then(data => {
      if (data.ok) loadComentarios(key);
      else alert('Error al borrar');
    })
    .catch(() => alert('Error de conexi\u00f3n con el servidor.'));
}
</script>
</body>
</html>"""

html_clean = html.encode('utf-8', errors='replace').decode('utf-8')

out = BASE_PATH + "index.html"
with open(out, 'w', encoding='utf-8') as f:
    f.write(html_clean)
print(f"OK: {len(html_clean):,} bytes -> {out}")
